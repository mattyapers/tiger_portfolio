"""Dump Holdings + Rebalance + Entry Signals from xlsx for review."""
import json
from openpyxl import load_workbook

import os
os.chdir(os.path.dirname(os.path.abspath(__file__)))
wb = load_workbook('output/portfolio_tracker.xlsx', data_only=False)

def dump_sheet(name, max_rows=40):
    if name not in wb.sheetnames:
        print(f"-- missing sheet: {name}")
        return
    ws = wb[name]
    print(f"\n===== {name} =====")
    for r in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows), values_only=True):
        if any(c is not None for c in r):
            print(r)

for s in ['Dashboard', 'Holdings', 'Rebalance Signals', 'Entry Signals']:
    dump_sheet(s)
