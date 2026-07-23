"""
load.py — Stage 3: Write analytics results to Excel.

DESIGN CHOICES:
- Data cells get ACTUAL VALUES (from Tiger API / calculations)
- Summary cells get EXCEL FORMULAS (so the sheet stays dynamic)
- Color coding: Blue text = input, Black = formula, Yellow bg = attention
- This means you can manually tweak a price in Excel and totals update

WHY NOT ALL FORMULAS:
The holdings data comes from Tiger API — it's authoritative.
But summary/signal cells benefit from formulas so you can
do "what-if" scenarios by changing prices manually.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

# Style constants
BLUE_F = Font(color='0000FF', name='Arial', size=10)
BLACK_F = Font(color='000000', name='Arial', size=10)
BOLD_F = Font(bold=True, name='Arial', size=10)
HEADER_F = Font(color='FFFFFF', bold=True, name='Arial', size=10)
TITLE_F = Font(color='4FC3F7', bold=True, name='Arial', size=14)
SECTION_F = Font(color='FFD54F', bold=True, name='Arial', size=11)
INPUT_BG = PatternFill('solid', fgColor='FFFDE7')
HEADER_BG = PatternFill('solid', fgColor='162447')
ALERT_BG = PatternFill('solid', fgColor='FFCDD2')
GOOD_BG = PatternFill('solid', fgColor='C8E6C9')
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)


def _style_header_row(ws, row, cols, headers):
    """Apply header styling to a row."""
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = HEADER_F
        cell.fill = HEADER_BG
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER


def _auto_width(ws, df_columns, min_width=10, max_width=22):
    """Set column widths based on content."""
    from openpyxl.utils import get_column_letter
    for i, col in enumerate(df_columns, 1):
        width = max(min_width, min(max_width, len(str(col)) + 4))
        ws.column_dimensions[get_column_letter(i)].width = width


def write_holdings_sheet(wb, holdings_df, analytics_timestamp):
    """
    Write the Holdings sheet — your source of truth.

    Each row = one position with:
    - Static data (tier, ticker, shares, cost) from Tiger API
    - Formulas for calculated fields (market value, P&L, weights)
    """
    ws = wb.active
    ws.title = '📈 Holdings'

    # Title — shows when the data is FROM, not when you ran the script
    ws.merge_cells('A1:N1')
    data_date = analytics_timestamp.strftime('%Y-%m-%d') if hasattr(analytics_timestamp, 'strftime') else str(analytics_timestamp)
    ws['A1'] = f"PORTFOLIO HOLDINGS — Data as of {data_date}"
    ws['A1'].font = TITLE_F

    # Headers (row 3)
    headers = ['Tier', 'Ticker', 'Name', 'Current $', 'Shares', 'Cost/Avg',
               'Mkt Value', 'Cost Basis', 'Unreal P&L', 'P&L %',
               'Wt% Port', 'Wt% Sat', 'Target %', 'Status']
    _style_header_row(ws, 3, 14, headers)

    col_widths = [11, 8, 22, 11, 10, 11, 12, 12, 12, 9, 9, 9, 9, 16]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data rows (start at row 4)
    data_start = 4
    for idx, row in holdings_df.iterrows():
        r = data_start + idx

        ws.cell(row=r, column=1, value=row['tier'])
        ws.cell(row=r, column=2, value=row['symbol']).font = BOLD_F
        ws.cell(row=r, column=3, value=row.get('name', ''))

        # Price (INPUT — blue, yellow bg)
        c = ws.cell(row=r, column=4, value=row['latest_price'])
        c.font = BLUE_F; c.fill = INPUT_BG; c.number_format = '$#,##0.00'

        # Shares (INPUT)
        c = ws.cell(row=r, column=5, value=row['shares'])
        c.font = BLUE_F; c.fill = INPUT_BG; c.number_format = '0.0000'

        # Cost/Avg (INPUT)
        c = ws.cell(row=r, column=6, value=row['avg_cost'])
        c.font = BLUE_F; c.fill = INPUT_BG; c.number_format = '$#,##0.00'

        # Market Value = Price * Shares (FORMULA)
        ws.cell(row=r, column=7).value = f'=D{r}*E{r}'
        ws.cell(row=r, column=7).number_format = '$#,##0.00'

        # Cost Basis = Shares * CostAvg (FORMULA)
        ws.cell(row=r, column=8).value = f'=E{r}*F{r}'
        ws.cell(row=r, column=8).number_format = '$#,##0.00'

        # Unrealized P&L (FORMULA)
        ws.cell(row=r, column=9).value = f'=G{r}-H{r}'
        ws.cell(row=r, column=9).number_format = '$#,##0.00'

        # P&L % (FORMULA)
        ws.cell(row=r, column=10).value = f'=IF(H{r}=0,0,I{r}/H{r})'
        ws.cell(row=r, column=10).number_format = '0.0%'

        # Weight % Portfolio (FORMULA — references total in row 2)
        ws.cell(row=r, column=11).value = f'=IF(G$2=0,0,G{r}/G$2)'
        ws.cell(row=r, column=11).number_format = '0.0%'

        # Weight % Satellite (FORMULA — only for satellite)
        if row['tier'] == 'Satellite':
            ws.cell(row=r, column=12).value = f'=IF(SUMIF(A$4:A$99,"Satellite",G$4:G$99)=0,0,G{r}/SUMIF(A$4:A$99,"Satellite",G$4:G$99))'
        else:
            ws.cell(row=r, column=12, value='-')
        ws.cell(row=r, column=12).number_format = '0.0%'

        # Target %
        target = row.get('target_pct', '-')
        if row['tier'] == 'Satellite' and isinstance(target, (int, float)):
            c = ws.cell(row=r, column=13, value=target)
            c.font = BLUE_F; c.fill = INPUT_BG; c.number_format = '0.0%'
        else:
            ws.cell(row=r, column=13, value='-')

        # Status
        ws.cell(row=r, column=14, value=row.get('status', ''))

        for c in range(1, 15):
            ws.cell(row=r, column=c).border = THIN_BORDER

    # Summary row at top (row 2) — FORMULAS
    data_end = data_start + len(holdings_df) - 1
    ws['F2'] = 'Total:'
    ws['F2'].font = SECTION_F
    ws.cell(row=2, column=7).value = f'=SUM(G{data_start}:G{data_end})'
    ws.cell(row=2, column=7).number_format = '$#,##0.00'
    ws.cell(row=2, column=9).value = f'=SUM(I{data_start}:I{data_end})'
    ws.cell(row=2, column=9).number_format = '$#,##0.00'

    ws.freeze_panes = 'A4'
    return ws


def write_rebalance_sheet(wb, rebalance_df):
    """Write the Rebalance Signals sheet."""
    ws = wb.create_sheet('⚖️ Rebalance Signals')

    ws.merge_cells('A1:K1')
    ws['A1'] = 'REBALANCE SIGNALS — Satellite Positions'
    ws['A1'].font = TITLE_F

    headers = ['#', 'Ticker', 'Mkt Value', 'Wt% Sat', 'Target %',
               'Drift $', 'Drift %', 'Signal', 'Action', 'Shares to Trade', 'Est Proceeds']
    _style_header_row(ws, 3, 11, headers)

    col_widths = [5, 8, 12, 10, 10, 12, 10, 16, 28, 14, 12]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    if rebalance_df.empty:
        return ws

    for idx, (_, row) in enumerate(rebalance_df.iterrows()):
        r = 4 + idx
        ws.cell(row=r, column=1, value=idx + 1)
        ws.cell(row=r, column=2, value=row['symbol']).font = BOLD_F
        ws.cell(row=r, column=3, value=row['market_value']).number_format = '$#,##0.00'
        ws.cell(row=r, column=4, value=row['wt_satellite']).number_format = '0.0%'
        ws.cell(row=r, column=5, value=row['target_pct']).number_format = '0.0%'
        ws.cell(row=r, column=6, value=row['drift_dollars']).number_format = '$#,##0.00'
        ws.cell(row=r, column=7, value=row['drift_pct']).number_format = '+0.0%;-0.0%'
        ws.cell(row=r, column=8, value=row['signal'])
        ws.cell(row=r, column=9, value=row['action'])
        ws.cell(row=r, column=10, value=row['shares_to_trade']).number_format = '0.0000'
        ws.cell(row=r, column=11, value=row['est_proceeds']).number_format = '$#,##0.00'

        # Color-code signal cells
        sig_cell = ws.cell(row=r, column=8)
        if '🚨' in str(row['signal']):
            sig_cell.fill = ALERT_BG
        elif '💡' in str(row['signal']):
            sig_cell.fill = GOOD_BG

        for c in range(1, 12):
            ws.cell(row=r, column=c).border = THIN_BORDER

    ws.freeze_panes = 'A4'
    return ws


def write_entry_signals_sheet(wb, entry_df):
    """Write the Entry/Exit Signals sheet."""
    ws = wb.create_sheet('🎯 Entry Signals')

    ws.merge_cells('A1:J1')
    ws['A1'] = 'ENTRY & EXIT SIGNALS — Valuation Triggers'
    ws['A1'].font = TITLE_F

    headers = ['#', 'Ticker', 'Current $', 'P/E', '5Y Avg P/E',
               'PE vs 5Y', 'P&L %', 'Cost Dist %', 'Entry Score', 'Signal']
    _style_header_row(ws, 3, 10, headers)

    col_widths = [5, 8, 11, 8, 11, 11, 9, 11, 11, 28]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    if entry_df.empty:
        return ws

    for idx, (_, row) in enumerate(entry_df.iterrows()):
        r = 4 + idx
        ws.cell(row=r, column=1, value=idx + 1)
        ws.cell(row=r, column=2, value=row['symbol']).font = BOLD_F
        ws.cell(row=r, column=3, value=row['latest_price']).number_format = '$#,##0.00'

        pe = row.get('pe_ttm')
        ws.cell(row=r, column=4, value=pe if pd.notna(pe) else 'N/A')
        if pd.notna(pe):
            ws.cell(row=r, column=4).number_format = '0.0'

        pe5y = row.get('pe_5y_avg')
        c = ws.cell(row=r, column=5, value=pe5y if pd.notna(pe5y) else 'N/A')
        if pd.notna(pe5y):
            c.font = BLUE_F; c.fill = INPUT_BG; c.number_format = '0.0'

        prem = row.get('pe_premium')
        ws.cell(row=r, column=6, value=prem if pd.notna(prem) else 'N/A')
        if pd.notna(prem):
            ws.cell(row=r, column=6).number_format = '+0.0%;-0.0%'

        ws.cell(row=r, column=7, value=row.get('pnl_pct', 0)).number_format = '0.0%'
        ws.cell(row=r, column=8, value=row.get('cost_distance', 0)).number_format = '+0.0%;-0.0%'

        score = row.get('entry_score')
        ws.cell(row=r, column=9, value=score if pd.notna(score) else 'N/A')

        sig = row.get('entry_signal', '')
        sig_cell = ws.cell(row=r, column=10, value=sig)
        if '🚨' in str(sig) or '🛑' in str(sig):
            sig_cell.fill = ALERT_BG
        elif '💡' in str(sig):
            sig_cell.fill = GOOD_BG

        for c in range(1, 11):
            ws.cell(row=r, column=c).border = THIN_BORDER

    ws.freeze_panes = 'A4'
    return ws


def write_dashboard_sheet(wb, summary, settings, analytics_timestamp):
    """Write the Dashboard summary sheet."""
    ws = wb.create_sheet('📊 Dashboard')
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16

    ws.merge_cells('A1:D1')
    data_date = analytics_timestamp.strftime('%Y-%m-%d') if hasattr(analytics_timestamp, 'strftime') else str(analytics_timestamp)
    ws['A1'] = f"PORTFOLIO DASHBOARD — Data as of {data_date} | {settings.MACRO_REGIME.get('quadrant', 'Quadrant Unknown')}"
    ws['A1'].font = TITLE_F

    # Date tracking row
    ws['A2'] = 'Report generated:'
    ws['B2'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    ws['C2'] = 'Next review:'
    # Calculate next review date from snapshot + review cycle
    from datetime import timedelta
    if hasattr(analytics_timestamp, 'strftime'):
        next_review = analytics_timestamp + timedelta(days=settings.REBALANCE_RULES['review_cycle_days'])
        ws['D2'] = next_review.strftime('%Y-%m-%d')
    else:
        ws['D2'] = 'TBD'
    ws['D2'].font = BLUE_F
    ws['D2'].fill = INPUT_BG

    # Metrics
    metrics = [
        ('Total Portfolio', summary['total_portfolio'], '$#,##0.00'),
        ('Core Weight', summary['core_pct'], '0.0%'),
        ('Core-Plus Weight', summary['coreplus_pct'], '0.0%'),
        ('Satellite Weight', summary['satellite_pct'], '0.0%'),
        ('Total P&L ($)', summary['total_pnl'], '$#,##0.00'),
        ('Total P&L (%)', summary['total_pnl_pct'], '0.0%'),
        ('Positions', summary['positions_count'], '0'),
    ]

    _style_header_row(ws, 3, 4, ['Metric', 'Value', 'Target', 'Status'])
    targets = [None, 0.68, 0.11, 0.21, None, None, None]

    for idx, ((label, val, fmt), target) in enumerate(zip(metrics, targets)):
        r = 4 + idx
        ws.cell(row=r, column=1, value=label)
        c = ws.cell(row=r, column=2, value=val)
        c.number_format = fmt
        if target is not None:
            ws.cell(row=r, column=3, value=target).number_format = '0%'

    # Tier drift section
    ws['A13'] = 'TIER HEALTH'
    ws['A13'].font = SECTION_F
    _style_header_row(ws, 14, 4, ['Tier', 'Actual %', 'Target %', 'Status'])
    for idx, td in enumerate(summary.get('tier_drift', [])):
        r = 15 + idx
        ws.cell(row=r, column=1, value=td['tier'])
        ws.cell(row=r, column=2, value=td['actual_pct']).number_format = '0.0%'
        ws.cell(row=r, column=3, value=td['target_pct']).number_format = '0%'
        ws.cell(row=r, column=4, value=td['status'])

    # Macro regime
    ws['A20'] = 'MACRO REGIME (Manual Update)'
    ws['A20'].font = SECTION_F
    for idx, (key, val) in enumerate(settings.MACRO_REGIME.items()):
        r = 21 + idx
        ws.cell(row=r, column=1, value=key.replace('_', ' ').title())
        if isinstance(val, (list, tuple)):
            val = '; '.join(str(x) for x in val)
        c = ws.cell(row=r, column=2, value=val)
        c.font = BLUE_F; c.fill = INPUT_BG

    return ws


def write_audit_sheet(wb, audit):
    """
    Write the freshness + price-drift audit to a dedicated sheet.

    Lets the weekly Claude review read a structured table instead of
    inspecting source files for staleness.
    """
    ws = wb.create_sheet('📋 Audit')
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14

    ws.merge_cells('A1:E1')
    ws['A1'] = 'DATA FRESHNESS AUDIT'
    ws['A1'].font = TITLE_F

    is_clean = audit.get('is_clean', True)
    ws['A2'] = 'Overall status:'
    ws['B2'] = '✅ ALL FRESH' if is_clean else '🚨 STALE ITEMS'
    ws['B2'].font = BOLD_F

    # Freshness table
    _style_header_row(ws, 4, 5, ['Item', 'Last Updated', 'Age (d)', 'Cadence (d)', 'Status'])
    for idx, r in enumerate(audit.get('freshness', [])):
        row = 5 + idx
        ws.cell(row=row, column=1, value=r['item'])
        ws.cell(row=row, column=2, value=r['last_updated'])
        ws.cell(row=row, column=3, value=r['age_days'])
        ws.cell(row=row, column=4, value=r['cadence_days'])
        c = ws.cell(row=row, column=5, value=r['status'])
        if 'STALE' in r['status']:
            c.fill = ALERT_BG
        else:
            c.fill = GOOD_BG
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER

    # Price drift section
    drift = audit.get('price_drift', [])
    drift_start = 5 + len(audit.get('freshness', [])) + 2
    ws.cell(row=drift_start, column=1, value='PRICE DRIFT vs SPOT (>10%)').font = SECTION_F

    if not drift:
        ws.cell(row=drift_start + 1, column=1, value='✅ No positions drifted >10% from snapshot')
    else:
        _style_header_row(ws, drift_start + 1, 4,
                          ['Symbol', 'Snapshot $', 'Spot $', 'Drift %'])
        for idx, f in enumerate(drift):
            row = drift_start + 2 + idx
            ws.cell(row=row, column=1, value=f['symbol']).font = BOLD_F
            ws.cell(row=row, column=2, value=f['snapshot_price']).number_format = '$#,##0.00'
            ws.cell(row=row, column=3, value=f['spot_price']).number_format = '$#,##0.00'
            c = ws.cell(row=row, column=4, value=f['drift_pct'])
            c.number_format = '+0.0%;-0.0%'
            c.fill = ALERT_BG
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = THIN_BORDER

    return ws


def write_watchlist_sheet(wb, settings):
    """Write the WATCHLIST dict to a structured sheet."""
    ws = wb.create_sheet('👀 Watchlist')

    ws.merge_cells('A1:F1')
    ws['A1'] = 'WATCHLIST — Pending Actions & Watch Items'
    ws['A1'].font = TITLE_F

    headers = ['Key', 'Ticker', 'Action', 'Target Price', 'Date', 'Note']
    _style_header_row(ws, 3, 6, headers)

    from openpyxl.utils import get_column_letter
    for i, w in enumerate([18, 8, 14, 13, 12, 60], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for idx, (key, entry) in enumerate(settings.WATCHLIST.items()):
        r = 4 + idx
        ws.cell(row=r, column=1, value=key)
        ws.cell(row=r, column=2, value=entry.get('ticker', '')).font = BOLD_F

        action = entry.get('action', '—')
        action_cell = ws.cell(row=r, column=3, value=action)
        if action in ('EXIT', 'TRIM 50%'):
            action_cell.fill = ALERT_BG
        elif action == 'WATCH':
            action_cell.fill = GOOD_BG

        target = entry.get('target_price')
        if target is not None:
            ws.cell(row=r, column=4, value=target).number_format = '$#,##0.00'
        else:
            ws.cell(row=r, column=4, value='—')

        date_val = (entry.get('trigger_date') or entry.get('review_date')
                    or entry.get('catalyst_date') or '—')
        ws.cell(row=r, column=5, value=date_val)

        note = entry.get('note', '')
        ws.cell(row=r, column=6, value=note[:200] if note else '').alignment = Alignment(wrap_text=True)

        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN_BORDER

    ws.freeze_panes = 'A4'
    return ws


def write_screener_sheet(wb, screener_df):
    """Write the watchlist valuation screener results."""
    ws = wb.create_sheet('🔍 Screener')

    ws.merge_cells('A1:J1')
    ws['A1'] = 'WATCHLIST SCREENER — Valuation + Regime Fit (Quadrant D)'
    ws['A1'].font = TITLE_F

    headers = ['Ticker', 'Price', 'P/E TTM', 'Fwd P/E', '5Y Avg P/E',
               'P/E Premium', 'FCF Yield', 'Regime Fit', 'Regime Note', 'Signal']
    _style_header_row(ws, 3, 10, headers)

    from openpyxl.utils import get_column_letter
    for i, w in enumerate([8, 10, 9, 9, 11, 11, 10, 12, 36, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    if screener_df.empty:
        ws.cell(row=4, column=1,
                value='No screener data — run in hybrid or yf-only mode')
        return ws

    for idx, (_, row) in enumerate(screener_df.iterrows()):
        r = 4 + idx
        ws.cell(row=r, column=1, value=row['symbol']).font = BOLD_F

        price = row.get('price')
        if pd.notna(price) and price is not None:
            ws.cell(row=r, column=2, value=price).number_format = '$#,##0.00'
        else:
            ws.cell(row=r, column=2, value='—')

        for col, field in [(3, 'pe_ttm'), (4, 'fwd_pe'), (5, 'pe_5y_avg')]:
            val = row.get(field)
            c = ws.cell(row=r, column=col, value=val if pd.notna(val) and val is not None else '—')
            if pd.notna(val) and val is not None:
                c.number_format = '0.0'

        prem = row.get('pe_premium_pct')
        if pd.notna(prem) and prem is not None:
            ws.cell(row=r, column=6, value=prem / 100).number_format = '+0.0%;-0.0%'
        else:
            ws.cell(row=r, column=6, value='—')

        fcfy = row.get('fcf_yield_pct')
        if pd.notna(fcfy) and fcfy is not None:
            ws.cell(row=r, column=7, value=fcfy / 100).number_format = '0.0%'
        else:
            ws.cell(row=r, column=7, value='—')

        fit = str(row.get('regime_fit', '—'))
        fit_cell = ws.cell(row=r, column=8, value=fit)
        if '✅' in fit:
            fit_cell.fill = GOOD_BG
        elif '❌' in fit:
            fit_cell.fill = ALERT_BG

        ws.cell(row=r, column=9, value=row.get('regime_note', ''))

        sig = str(row.get('signal', ''))
        sig_cell = ws.cell(row=r, column=10, value=sig)
        if '🚨' in sig:
            sig_cell.fill = ALERT_BG
        elif '💡' in sig:
            sig_cell.fill = GOOD_BG

        for c in range(1, 11):
            ws.cell(row=r, column=c).border = THIN_BORDER

    ws.freeze_panes = 'A4'
    return ws


def load_to_excel(analytics, settings):
    """
    Master load function — writes all sheets to Excel.

    INPUT: analytics dict from transform stage
    OUTPUT: Excel file at settings.OUTPUT_PATH
    """
    wb = Workbook()

    # Enrich holdings with targets for Excel output
    holdings = analytics['holdings'].copy()
    holdings['target_pct'] = holdings['symbol'].map(settings.SATELLITE_TARGETS)
    holdings['status'] = ''
    for _, row in analytics['rebalance'].iterrows():
        mask = holdings['symbol'] == row['symbol']
        holdings.loc[mask, 'status'] = row['signal']
    # Core positions get status
    holdings.loc[holdings['tier'].isin(['Core', 'Core-Bond', 'Core-Plus']), 'status'] = '✅ Core'

    # Add name column
    name_map = {
        'VXUS': 'Vanguard Total Intl', 'VOO': 'Vanguard S&P 500',
        'BND': 'Vanguard Total Bond', 'IEF': 'iShares 7-10Y Treasury',
        'SPTL': 'SPDR Long Treasury', 'SHY': 'iShares 1-3Y Treasury',
        'VTIP': 'Vanguard Short TIPS', 'SPYD': 'SPDR Div ETF',
        'ONEQ': 'Fidelity Nasdaq', 'GLDM': 'SPDR Gold Mini',
        'GOOG': 'Alphabet', 'RTX': 'RTX Corp', 'NVDA': 'NVIDIA',
        'TSM': 'Taiwan Semi', 'AAPL': 'Apple', 'MA': 'MasterCard',
        'CAT': 'Caterpillar', 'KO': 'Coca-Cola', 'BABA': 'Alibaba',
        'AON': 'Aon PLC', 'XLE': 'Energy Select SPDR',
        'MSFT': 'Microsoft',
        'COP':  'ConocoPhillips',
        'NTR':  'Nutrien',
    }
    holdings['name'] = holdings['symbol'].map(name_map).fillna('')

    # Reset index for clean row numbering
    holdings = holdings.reset_index(drop=True)

    write_holdings_sheet(wb, holdings, analytics['timestamp'])
    write_rebalance_sheet(wb, analytics['rebalance'])
    write_entry_signals_sheet(wb, analytics['entry_signals'])
    write_dashboard_sheet(wb, analytics['summary'], settings, analytics['timestamp'])
    if 'audit' in analytics:
        write_audit_sheet(wb, analytics['audit'])
    write_watchlist_sheet(wb, settings)
    write_screener_sheet(wb, analytics.get('screener', pd.DataFrame()))

    # Move Dashboard to first position (7 sheets total: Holdings + 6 others)
    wb.move_sheet('📊 Dashboard', offset=-6)

    wb.save(settings.OUTPUT_PATH)
    logger.info(f"Excel saved to {settings.OUTPUT_PATH}")
    return settings.OUTPUT_PATH